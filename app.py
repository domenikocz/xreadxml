import streamlit as st
import pandas as pd
import xml.etree.ElementTree as ET
import io
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# --- FUNZIONI DI LOGICA (Traduzione Macro VBA) ---

def determina_tipo_documento(tipo_doc):
    mapping = {
        "TD01": "FATTURA", "TD02": "FATTURA", "TD03": "FATTURA",
        "TD04": "NOTA CREDITO", "TD05": "NOTA DEBITO", "TD06": "PARCELLA"
    }
    return mapping.get(tipo_doc, "ALTRO")

def processa_xml(file):
    tree = ET.parse(file)
    root = tree.getroot()
    # Rimuove i namespace per facilitare la ricerca dei tag
    for elem in root.iter():
        if '}' in elem.tag:
            elem.tag = elem.tag.split('}', 1)[1]

    dati = {}
    
    # Estrazione Dati Fornitore
    cp = root.find(".//CedentePrestatore/DatiAnagrafici")
    if cp is not None:
        dati['P.IVA'] = cp.findtext(".//IdCodice", "")
        dati['COD.FISC'] = cp.findtext("CodiceFiscale", "")
        denominazione = cp.findtext(".//Denominazione", "")
        if not denominazione:
            denominazione = f"{cp.findtext('.//Nome', '')} {cp.findtext('.//Cognome', '')}".strip()
        dati['DENOMINAZIONE'] = denominazione

    # Estrazione Dati Generali
    dg = root.find(".//DatiGeneraliDocumento")
    if dg is not None:
        dati['NUMERO'] = dg.findtext("Numero", "")
        data_str = dg.findtext("Data", "")
        dati['DATA'] = datetime.strptime(data_str, '%Y-%m-%d').date() if data_str else None
        dati['TIPO'] = determina_tipo_documento(dg.findtext("TipoDocumento", ""))
        dati['BOLLO'] = float(dg.findtext("ImportoBollo", "0").replace(",", "."))
        dati['TOTALE'] = float(dg.findtext("ImportoTotaleDocumento", "0").replace(",", "."))
        dati['CAUSALE'] = dg.findtext("Causale", "")

    # Somma Imponibile e IVA (come da macro VBA)
    tot_imponibile = 0.0
    tot_iva = 0.0
    for riepilogo in root.findall(".//DatiRiepilogo"):
        imp = riepilogo.findtext("ImponibileImporto", "0").replace(",", ".")
        iva = riepilogo.findtext("Imposta", "0").replace(",", ".")
        tot_imponibile += float(imp)
        tot_iva += float(iva)
    
    dati['IMPONIBILE'] = tot_imponibile
    dati['IVA'] = tot_iva

    # Ritenute
    dr = root.find(".//DatiRitenuta")
    dati['RITENUTE'] = float(dr.findtext("ImportoRitenuta", "0").replace(",", ".")) if dr is not None else 0.0
    dati['TIPO RIT.'] = dr.findtext("TipoRitenuta", "") if dr is not None else ""

    # Descrizione Linee
    descrizioni = [linea.text for linea in root.findall(".//DettaglioLinee/Descrizione") if linea.text]
    full_desc = "; ".join(descrizioni)
    dati['DESCRIZIONE'] = (full_desc[:252] + '...') if len(full_desc) > 255 else full_desc
    dati['NOME FILE'] = file.name
    return dati

def esporta_excel_formattato(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Fatture')
        workbook = writer.book
        worksheet = writer.sheets['Fatture']

        # Definizione Stili (Colori Originali Macro)
        header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        font_bold = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        color_map = {
            "FATTURA": Font(color="008000", bold=True),      # Verde
            "PARCELLA": Font(color="800080", bold=True),     # Viola
            "NOTA CREDITO": Font(color="800000", bold=True)  # Rosso/Amaranto
        }

        # Blocca Riquadri
        worksheet.freeze_panes = "A2"

        # Formattazione Intestazioni
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = font_bold
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Formattazione Dati
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1):
            for cell in row:
                cell.border = thin_border
                # Formato Valuta per colonne 6-10 (Imponibile, IVA, Totale, Ritenute, Bollo)
                if cell.column in range(6, 11): 
                    cell.number_format = '#,##0.00 "€"'
                # Colore condizionale per TIPO
                if cell.column == 5:
                    valore_tipo = str(cell.value).upper()
                    if valore_tipo in color_map:
                        cell.font = color_map[valore_tipo]
            # Formato Data
            worksheet.cell(row=row[0].row, column=4).number_format = 'DD/MM/YYYY'

        # Auto-fit Colonne
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            worksheet.column_dimensions[column_letter].width = max_length + 2

    return output.getvalue()

# --- INTERFACCIA WEB STREAMLIT ---

st.set_page_config(page_title="Convertitore XML", layout="wide")
st.title("📂 Elaborazione Fatture XML")

uploaded_files = st.file_uploader("Trascina qui i tuoi file XML", type="xml", accept_multiple_files=True)

if uploaded_files:
    lista_finale = []
    visti = set()

    for f in uploaded_files:
        try:
            d = processa_xml(f)
            # Controllo duplicati: P.IVA + Numero + Anno (come da Macro)
            anno_fatt = d['DATA'].year if d['DATA'] else ""
            chiave = (d['P.IVA'], d['NUMERO'], anno_fatt)
            
            if chiave not in visti:
                lista_finale.append(d)
                visti.add(chiave)
            else:
                st.warning(f"Duplicato ignorato: {d['NOME FILE']} (P.IVA {d['P.IVA']} - Num. {d['NUMERO']})")
        except Exception as e:
            st.error(f"Errore nell'elaborazione di {f.name}: {e}")

    if lista_finale:
        df = pd.DataFrame(lista_finale)
        # Ordine colonne come da richiesta originale
        cols = ["P.IVA", "DENOMINAZIONE", "NUMERO", "DATA", "TIPO", "IMPONIBILE", "IVA", "TOTALE", "RITENUTE", "BOLLO", "TIPO RIT.", "CAUSALE", "COD.FISC", "DESCRIZIONE", "NOME FILE"]
        df = df[cols]
        
        st.subheader("Anteprima Elaborazione")
        st.dataframe(df)
        
        # Generazione file Excel formattato
        excel_data = esporta_excel_formattato(df)
        
        st.download_button(
            label="📥 Scarica Excel Formattato",
            data=excel_data,
            file_name="estrazione_fatture.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
